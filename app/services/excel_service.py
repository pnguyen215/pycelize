"""
Excel Service Implementation

This module provides Excel file processing functionality including
reading, writing, column extraction, and data manipulation.
"""

import os
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from app.core.config import Config
from app.core.exceptions import FileProcessingError, ValidationError
from app.core.logging import get_logger

logger = get_logger(__name__)


class ExcelService:
    """
    Service for Excel file processing operations.

    This service provides comprehensive Excel file handling including:
    - Reading Excel files
    - Writing DataFrames to Excel
    - Column extraction with deduplication
    - Column mapping and transformation
    - Auto-adjusting column widths

    Attributes:
        config: Application configuration
        max_column_width: Maximum column width for auto-adjustment
        default_sheet_name: Default name for data sheets

    Example:
        >>> service = ExcelService(config)
        >>> df = service.read_excel('data.xlsx')
        >>> columns = service.extract_columns(df, ['name', 'email'])
    """

    def __init__(self, config: Config):
        """
        Initialize ExcelService with configuration.

        Args:
            config: Application configuration instance
        """
        self.config = config
        self.max_column_width = config.get("excel.max_column_width", 50)
        self.default_sheet_name = config.get("excel.default_sheet_name", "Sheet1")
        self.include_info_sheet = config.get("excel.include_info_sheet", True)
        self.output_folder = config.get("file.output_folder", "outputs")

        # Ensure output folder exists
        os.makedirs(self.output_folder, exist_ok=True)

    def read_excel(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Read an Excel file into a DataFrame.

        Args:
            file_path: Path to the Excel file
            sheet_name: Optional specific sheet name to read

        Returns:
            DataFrame containing the Excel data

        Raises:
            FileProcessingError: If file cannot be read
        """
        try:
            logger.info(f"Reading Excel file: {file_path}")

            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)

            logger.info(f"Successfully read {len(df)} rows, {len(df.columns)} columns")
            return df

        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise FileProcessingError(f"Failed to read Excel file: {str(e)}")

    def write_excel(
        self,
        data: pd.DataFrame,
        output_path: str,
        sheet_name: Optional[str] = None,
        include_info: bool = True,
        auto_adjust: bool = True,
    ) -> str:
        """
        Write a DataFrame to an Excel file.

        Args:
            data: DataFrame to write
            output_path: Path for the output file
            sheet_name: Name for the data sheet
            include_info: Whether to include an info sheet
            auto_adjust: Whether to auto-adjust column widths

        Returns:
            Path to the created file

        Raises:
            FileProcessingError: If file cannot be written
        """
        try:
            sheet_name = sheet_name or self.default_sheet_name
            logger.info(f"Writing Excel file: {output_path}")

            workbook = Workbook()

            # Remove default sheet
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            # Create data sheet
            self._create_data_sheet(workbook, data, sheet_name, auto_adjust)

            # Create info sheet if requested
            if include_info and self.include_info_sheet:
                self._create_info_sheet(workbook, data, output_path)

            # Save workbook
            workbook.save(output_path)
            logger.info(f"Successfully wrote Excel file: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error writing Excel file: {str(e)}")
            raise FileProcessingError(f"Failed to write Excel file: {str(e)}")

    def extract_columns(
        self,
        data: pd.DataFrame,
        columns: List[str],
        remove_duplicates: bool = False,
        include_statistics: bool = True,
    ) -> Dict[str, Any]:
        """
        Extract data from specific columns.

        Args:
            data: Source DataFrame
            columns: List of column names to extract
            remove_duplicates: Whether to remove duplicate values
            include_statistics: Whether to include column statistics

        Returns:
            Dictionary containing extracted column data and statistics

        Raises:
            ValidationError: If column doesn't exist
        """
        logger.info(f"Extracting columns: {columns}")

        result = {
            "extraction_info": {
                "total_rows": len(data),
                "extracted_columns": columns,
                "remove_duplicates": remove_duplicates,
                "timestamp": datetime.now().isoformat(),
            },
            "columns": {},
        }

        for column in columns:
            if column not in data.columns:
                logger.warning(f"Column '{column}' not found in data")
                continue

            column_data = data[column]

            # Get values (optionally unique)
            if remove_duplicates:
                values = column_data.dropna().unique().tolist()
            else:
                values = column_data.tolist()

            column_info = {"values": values, "value_count": len(values)}

            # Add statistics if requested
            if include_statistics:
                column_info.update(
                    {
                        "total_count": len(column_data),
                        "non_null_count": int(column_data.notna().sum()),
                        "null_count": int(column_data.isna().sum()),
                        "unique_count": int(column_data.nunique()),
                        "data_type": str(column_data.dtype),
                    }
                )

                # Add numeric statistics if applicable
                if pd.api.types.is_numeric_dtype(column_data):
                    column_info["statistics"] = {
                        "mean": (
                            float(column_data.mean()) if not column_data.empty else None
                        ),
                        "median": (
                            float(column_data.median())
                            if not column_data.empty
                            else None
                        ),
                        "min": (
                            float(column_data.min()) if not column_data.empty else None
                        ),
                        "max": (
                            float(column_data.max()) if not column_data.empty else None
                        ),
                        "std": (
                            float(column_data.std()) if not column_data.empty else None
                        ),
                    }

            result["columns"][column] = column_info
            logger.debug(f"Extracted column '{column}': {len(values)} values")

        return result

    def apply_column_mapping(
        self, data: pd.DataFrame, mapping: Dict[str, Any]
    ) -> pd.DataFrame:
        """
        Apply column mapping to create new DataFrame with mapped columns.

        Args:
            data: Source DataFrame
            mapping: Dictionary mapping new column names to source columns or values
                    Format: {
                        'new_col': 'source_col',  # Simple mapping
                        'new_col': {'source': 'source_col', 'default': value},  # With default
                        'new_col': {'default': value}  # Default value only
                    }

        Returns:
            New DataFrame with mapped columns
        """
        logger.info(f"Applying column mapping with {len(mapping)} rules")

        mapped_data = {}

        for target_col, rule in mapping.items():
            if isinstance(rule, str):
                # Simple mapping: target <- source
                if rule in data.columns:
                    mapped_data[target_col] = data[rule]
                    logger.debug(f"Mapped '{target_col}' <- '{rule}'")
                else:
                    logger.warning(f"Source column '{rule}' not found")
                    mapped_data[target_col] = pd.Series([None] * len(data))

            elif isinstance(rule, dict):
                source = rule.get("source")
                default = rule.get("default")

                if source and source in data.columns:
                    series = data[source]
                    if default is not None:
                        series = series.fillna(default)
                    mapped_data[target_col] = series
                else:
                    mapped_data[target_col] = pd.Series([default] * len(data))

            else:
                # Create column with None values
                mapped_data[target_col] = pd.Series([None] * len(data))

        result = pd.DataFrame(mapped_data)
        logger.info(f"Created mapped DataFrame with {len(result.columns)} columns")

        return result

    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get sheet names from an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of sheet names
        """
        try:
            return pd.ExcelFile(file_path).sheet_names
        except Exception as e:
            logger.error(f"Error reading sheet names: {str(e)}")
            raise FileProcessingError(f"Failed to read sheet names: {str(e)}")

    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        Get information about an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary containing file information
        """
        try:
            df = self.read_excel(file_path)
            sheet_names = self.get_sheet_names(file_path)

            return {
                "file_path": file_path,
                "file_name": os.path.basename(file_path),
                "sheets": sheet_names,
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": list(df.columns),
                "data_types": {col: str(dtype) for col, dtype in df.dtypes.items()},
            }
        except Exception as e:
            raise FileProcessingError(f"Failed to get file info: {str(e)}")

    def _create_data_sheet(
        self, workbook: Workbook, data: pd.DataFrame, sheet_name: str, auto_adjust: bool
    ) -> None:
        """Create a data sheet in the workbook."""
        worksheet = workbook.create_sheet(title=sheet_name)

        # Write data
        for row in dataframe_to_rows(data, index=False, header=True):
            worksheet.append(row)

        # Auto-adjust column widths
        if auto_adjust:
            self._auto_adjust_column_width(worksheet, data)

    def _create_info_sheet(
        self, workbook: Workbook, data: pd.DataFrame, output_path: str
    ) -> None:
        """Create an info sheet with metadata."""
        info_data = [
            ["Property", "Value"],
            ["File", os.path.basename(output_path)],
            ["Created", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Rows", len(data)],
            ["Columns", len(data.columns)],
            ["Column Names", ", ".join(data.columns.tolist())],
        ]

        worksheet = workbook.create_sheet(title="Info")
        for row in info_data:
            worksheet.append(row)

    def _auto_adjust_column_width(self, worksheet, data: pd.DataFrame) -> None:
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, self.max_column_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def extract_columns_to_file(
        self,
        data: pd.DataFrame,
        columns: List[str],
        output_path: str,
        remove_duplicates: bool = False,
    ) -> str:
        """
        Extract specific columns from DataFrame and save to Excel file.

        Args:
            data: Source DataFrame
            columns: List of column names to extract
            output_path: Path for the output Excel file
            remove_duplicates: Whether to remove duplicate rows in extracted data

        Returns:
            Path to the created file

        Raises:
            ValidationError: If column doesn't exist
            FileProcessingError: If file cannot be written
        """
        logger.info(f"Extracting columns {columns} to file: {output_path}")

        # Validate columns
        missing_columns = [col for col in columns if col not in data.columns]
        if missing_columns:
            raise ValidationError(
                f"Columns not found in data: {', '.join(missing_columns)}"
            )

        # Extract columns
        extracted_df = data[columns].copy()

        # Remove duplicates if requested
        if remove_duplicates:
            original_count = len(extracted_df)
            extracted_df = extracted_df.drop_duplicates()
            logger.info(f"Removed {original_count - len(extracted_df)} duplicate rows")

        # Write to Excel
        self.write_excel(
            extracted_df,
            output_path,
            sheet_name="Sheet1",
            include_info=True,
            auto_adjust=True,
        )

        logger.info(
            f"Successfully extracted {len(columns)} columns with {len(extracted_df)} rows"
        )
        return output_path
