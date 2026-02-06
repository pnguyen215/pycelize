"""
Tests for Excel Service - Read Operations

This module contains unit tests specifically for ExcelService read operations,
including tests for preserving leading zeros and raw cell values.
"""

import os
import pytest
import pandas as pd
import tempfile
from openpyxl import Workbook

from app.core.config import Config
from app.services.excel_service import ExcelService


class TestExcelServiceRead:
    """Test cases for ExcelService read operations."""

    @pytest.fixture
    def config(self):
        """Create a test configuration."""
        config_content = """
app:
  name: "Test"
  version: "v0.0.1"

file:
  upload_folder: "test_uploads"
  output_folder: "test_outputs"

excel:
  default_sheet_name: "Sheet1"
  max_column_width: 50
  include_info_sheet: false
"""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".yml", delete=False) as f:
            f.write(config_content)
            config_path = f.name

        config = Config(config_path)
        yield config
        os.unlink(config_path)

    @pytest.fixture
    def service(self, config):
        """Create ExcelService instance."""
        return ExcelService(config)

    @pytest.fixture
    def excel_with_leading_zeros(self, tmp_path):
        """Create an Excel file with leading zeros for testing."""
        excel_path = tmp_path / "test_leading_zeros.xlsx"
        
        # Create workbook with leading zeros
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        
        # Add headers
        ws.append(['zip_code', 'phone', 'product_code', 'customer_id'])
        
        # Add data with leading zeros - store as text
        ws.append(['021201', '0123456789', '000456', '00001234'])
        ws.append(['00123', '555-1234', 'ABC123', '00005678'])
        ws.append(['12345', '999-8888', '789', '12345678'])
        
        # Format all cells as text to preserve leading zeros
        for row in range(2, 5):
            for col in range(1, 5):
                ws.cell(row, col).number_format = '@'  # Text format
        
        wb.save(excel_path)
        return str(excel_path)

    def test_read_excel_preserves_leading_zeros(self, service, excel_with_leading_zeros):
        """Test that reading Excel preserves leading zeros."""
        df = service.read_excel(excel_with_leading_zeros)
        
        # Verify leading zeros are preserved
        assert df['zip_code'].iloc[0] == '021201'
        assert df['zip_code'].iloc[1] == '00123'
        assert df['phone'].iloc[0] == '0123456789'
        assert df['product_code'].iloc[0] == '000456'
        assert df['customer_id'].iloc[0] == '00001234'
        
        # Verify all data is string type (either 'object' or pd.StringDtype())
        for col in ['zip_code', 'phone', 'product_code', 'customer_id']:
            assert df[col].dtype == 'object' or isinstance(df[col].dtype, pd.StringDtype)

    def test_read_excel_preserves_mixed_values(self, service, tmp_path):
        """Test that reading Excel preserves mixed alphanumeric values."""
        excel_path = tmp_path / "test_mixed.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.append(['code', 'description'])
        ws.append(['0001ABC', 'Product 1'])
        ws.append(['XYZ0002', 'Product 2'])
        ws.append(['00000', 'Product 3'])
        
        # Format as text
        for row in range(2, 5):
            for col in range(1, 3):
                ws.cell(row, col).number_format = '@'
        
        wb.save(excel_path)
        
        df = service.read_excel(str(excel_path))
        
        # Verify mixed values preserved
        assert df['code'].iloc[0] == '0001ABC'
        assert df['code'].iloc[1] == 'XYZ0002'
        assert df['code'].iloc[2] == '00000'

    def test_read_excel_with_sheet_name(self, service, tmp_path):
        """Test reading specific sheet with leading zeros preserved."""
        excel_path = tmp_path / "test_sheets.xlsx"
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(['id', 'value'])
        ws1.append(['001', 'A'])
        ws1.cell(2, 1).number_format = '@'
        
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(['code', 'name'])
        ws2.append(['002', 'B'])
        ws2.cell(2, 1).number_format = '@'
        
        wb.save(excel_path)
        
        df = service.read_excel(str(excel_path), sheet_name="Sheet2")
        
        assert df['code'].iloc[0] == '002'
        assert df['code'].dtype == 'object' or isinstance(df['code'].dtype, pd.StringDtype)
